from __future__ import annotations

import io
import os
import json
from contextlib import contextmanager
from datetime import datetime, date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

import psycopg
from psycopg.rows import dict_row

from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


# -----------------------------
# Connection / bootstrap
# -----------------------------

def _db_url() -> str:
    url = (os.getenv("DATABASE_URL") or os.getenv("POSTGRES_URL") or "").strip()
    if not url:
        raise RuntimeError("DATABASE_URL is not set")
    return url


@contextmanager
def get_conn():
    with psycopg.connect(_db_url(), row_factory=dict_row) as conn:
        yield conn


def init_db() -> None:
    """
    Idempotent schema init for Postgres.
    Safe to run on every startup.
    """
    with get_conn() as conn:
        # Users
        conn.execute(
            """
            create table if not exists users (
              id bigserial primary key,
              username text not null unique,
              created_at timestamptz not null default now()
            );
            """
        )

        # Schema meta (simple migration versioning)
        conn.execute(
            """
            create table if not exists schema_meta (
              id int primary key default 1,
              version int not null default 0,
              updated_at timestamptz not null default now()
            );
            insert into schema_meta(id, version) values (1, 0)
            on conflict (id) do nothing;
            """
        )

        # Audit log
        conn.execute(
            """
            create table if not exists audit_log (
              id bigserial primary key,
              user_id bigint,
              username text,
              action text not null,
              offer_id bigint,
              ip text,
              meta jsonb,
              created_at timestamptz not null default now()
            );
            """
        )

        # Offers (legacy user_name retained; v2 uses user_id)
        conn.execute(
            """
            create table if not exists offers (
              id bigserial primary key,
              user_id bigint,
              user_name text not null,
              client_name text,
              client_email text,
              client_address text,
              client_oib text,
              created_at timestamptz not null default now(),
              sent_at timestamptz,
              last_email_to text,
              last_email_at timestamptz,
              last_email_error text,
              email_attempts int not null default 0,


              -- numbering
              offer_year int,
              offer_seq int,
              offer_no text,

              -- workflow
              status text not null default 'DRAFT',
              accepted_at timestamptz,

              -- archive (soft delete)
              archived boolean not null default false,
              archived_at timestamptz,

              -- meta for PDF
              terms_delivery text,
              terms_payment text,
              note text,
              place text,
              signed_by text,
              vat_rate double precision not null default 0
            );
            """
        )

        
        # Offers - portal/tracking fields (idempotent)
        conn.execute("alter table offers add column if not exists public_token text;")
        conn.execute("alter table offers add column if not exists view_count int not null default 0;")
        conn.execute("alter table offers add column if not exists first_view_at timestamptz;")
        conn.execute("alter table offers add column if not exists last_view_at timestamptz;")
        conn.execute("alter table offers add column if not exists last_view_ip text;")
        conn.execute("alter table offers add column if not exists click_count int not null default 0;")
        conn.execute("alter table offers add column if not exists last_click_at timestamptz;")
        conn.execute("alter table offers add column if not exists last_click_ip text;")
        conn.execute("alter table offers add column if not exists accepted_via text;")
        conn.execute("alter table offers add column if not exists accepted_via_at timestamptz;")
        conn.execute("create unique index if not exists offers_public_token_uix on offers(public_token) where public_token is not null;")
        conn.execute("alter table offers add column if not exists valid_until date;")
# Items
        conn.execute(
            """
            create table if not exists offer_items (
              id bigserial primary key,
              offer_id bigint not null references offers(id) on delete cascade,
              name text not null,
              qty double precision not null default 1,
              price double precision not null default 0,
              line_total double precision not null default 0,
              created_at timestamptz not null default now()
            );
            """
        )

        # Company settings (legacy user_name PK retained; add user_id + logo bytes)
        conn.execute(
            """
            create table if not exists company_settings (
              user_name text primary key,
              company_name text,
              company_address text,
              company_oib text,
              company_iban text,
              company_email text,
              company_phone text,
              logo_path text,
              created_at timestamptz not null default now(),
              updated_at timestamptz not null default now()
            );
            """
        )

        # Clients
        conn.execute(
            """
            create table if not exists clients (
              id bigserial primary key,
              user_id bigint,
              user_name text not null,
              name text not null,
              email text,
              address text,
              oib text,
              note text,
              created_at timestamptz not null default now()
            );
            """
        )

        # --- Migrations (ADD COLUMN IF NOT EXISTS) ---
        # Offers: user_id + archive fields (for older DBs)
        conn.execute("alter table offers add column if not exists user_id bigint;")
        conn.execute("alter table offers add column if not exists archived boolean not null default false;")
        conn.execute("alter table offers add column if not exists archived_at timestamptz;")

        # Offers: status workflow extras
        conn.execute("alter table offers add column if not exists client_email text;")
        conn.execute("alter table offers add column if not exists sent_at timestamptz;")
        conn.execute("alter table offers add column if not exists last_email_to text;")
        conn.execute("alter table offers add column if not exists last_email_at timestamptz;")
        conn.execute("alter table offers add column if not exists last_email_error text;")
        conn.execute("alter table offers add column if not exists email_attempts int not null default 0;")

        # Offers: invoice fields
        conn.execute("alter table offers add column if not exists is_invoice boolean not null default false;")
        conn.execute("alter table offers add column if not exists invoice_year int;")
        conn.execute("alter table offers add column if not exists invoice_seq int;")
        conn.execute("alter table offers add column if not exists invoice_no text;")
        conn.execute("alter table offers add column if not exists invoice_date timestamptz;")
        conn.execute("alter table offers add column if not exists paid boolean not null default false;")
        conn.execute("alter table offers add column if not exists paid_at timestamptz;")

        # Company settings: templates
        conn.execute("alter table company_settings add column if not exists email_subject_tpl text;")
        conn.execute("alter table company_settings add column if not exists email_html_tpl text;")
        conn.execute("alter table company_settings add column if not exists email_text_tpl text;")
        conn.execute("alter table company_settings add column if not exists pdf_footer_tpl text;")


        # Settings: user_id + logo bytes
        conn.execute("alter table company_settings add column if not exists user_id bigint;")
        conn.execute("alter table company_settings add column if not exists logo_bytes bytea;")
        conn.execute("alter table company_settings add column if not exists logo_mime text;")
        conn.execute("alter table company_settings add column if not exists logo_filename text;")

        # Clients: user_id
        conn.execute("alter table clients add column if not exists user_id bigint;")

        # Clients: extra fields
        conn.execute("alter table clients add column if not exists email text;")
        conn.execute("alter table clients add column if not exists address text;")
        conn.execute("alter table clients add column if not exists oib text;")
        conn.execute("alter table clients add column if not exists note text;")

        # Offers: client extra fields
        conn.execute("alter table offers add column if not exists client_address text;")
        conn.execute("alter table offers add column if not exists client_oib text;")


        # Indexes (keep legacy indexes too)
        conn.execute("create index if not exists idx_offers_user_name on offers(user_name);")
        conn.execute("create index if not exists idx_offers_user_id on offers(user_id);")
        conn.execute("create index if not exists idx_offer_items_offer_id on offer_items(offer_id);")
        conn.execute("create unique index if not exists idx_clients_user_name_name on clients(user_name, name);")
        conn.execute("create unique index if not exists idx_users_username on users(username);")
        conn.execute("create unique index if not exists idx_company_settings_user_id on company_settings(user_id) where user_id is not null;")
        conn.execute("create index if not exists idx_clients_user_id on clients(user_id);")
        conn.execute("create index if not exists idx_offers_user_year_seq on offers(user_name, offer_year, offer_seq);")
        conn.execute("create index if not exists idx_offers_userid_year_seq on offers(user_id, offer_year, offer_seq);")
        conn.execute("create index if not exists idx_offers_invoice_user_year_seq on offers(user_id, invoice_year, invoice_seq);")

        # Backfill users from legacy tables (if any)
        conn.execute(
            """
            insert into users(username)
            select distinct lower(user_name) from (
              select user_name from offers
              union all
              select user_name from company_settings
              union all
              select user_name from clients
            ) s
            where user_name is not null and btrim(user_name) <> ''
            on conflict do nothing;
            """
        )

        # Backfill user_id columns using users table
        conn.execute(
            """
            update offers o
            set user_id = u.id
            from users u
            where o.user_id is null
              and lower(o.user_name) = u.username;
            """
        )
        conn.execute(
            """
            update clients c
            set user_id = u.id
            from users u
            where c.user_id is null
              and lower(c.user_name) = u.username;
            """
        )
        conn.execute(
            """
            update company_settings s
            set user_id = u.id
            from users u
            where s.user_id is null
              and lower(s.user_name) = u.username;
            """
        )

        # Backfill / ensure defaults
        conn.execute("update offers set vat_rate=0 where vat_rate is null;")
        conn.execute("update offers set status='DRAFT' where status is null;")
        try:
            conn.execute("alter table offers alter column status set default 'DRAFT';")
        except Exception:
            pass
        try:
            conn.execute("alter table offers alter column status set not null;")
        except Exception:
            pass


# -----------------------------
# Users
# -----------------------------

def ensure_user(username: str) -> int:
    username = (username or "").strip().lower()
    if not username:
        raise ValueError("username required")
    with get_conn() as conn:
        row = conn.execute("select id from users where username=%s", (username,)).fetchone()
        if row:
            return int(row["id"])
        row2 = conn.execute(
            "insert into users(username) values (%s) on conflict (username) do update set username=excluded.username returning id",
            (username,),
        ).fetchone()
        return int(row2["id"])


# -----------------------------
# Offers
# -----------------------------

def _offer_owner_clause() -> str:
    # Robust for legacy rows where user_id might be null.
    return "(o.user_id = %s or (o.user_id is null and lower(o.user_name) = %s))"



def next_offer_no(user_id: int, username: str, year: int | None = None) -> str:
    """Generate next offer number in format YYYY-0001 (per user)."""
    import datetime as _dt
    y = year or _dt.datetime.now().year
    prefix = f"{y}-"
    with get_conn() as conn:
        # Prefer already padded format for ordering
        row = conn.execute(
            "select offer_no from offers where (user_id=%s or user_name=%s) and (%s is null or client_id=%s) and offer_no like %s order by offer_no desc limit 1",
            (user_id, username, prefix + "%"),
        ).fetchone()
    if row and row.get("offer_no") and str(row["offer_no"]).startswith(prefix):
        tail = str(row["offer_no"])[len(prefix):]
        try:
            n = int(tail)
        except Exception:
            n = 0
        n += 1
    else:
        n = 1
    return f"{prefix}{n:04d}"

def create_offer(user_id: int, username: str, client_name: str | None = None) -> int:
    year = datetime.now().year
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        row_seq = conn.execute(
            """
            select coalesce(max(offer_seq), 0) as max_seq
            from offers
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
              and offer_year=%s
              and archived=false
            """,
            (user_id, username_l, year),
        ).fetchone()
        next_seq = int((row_seq or {}).get("max_seq") or 0) + 1
        offer_no = f"{year}-{next_seq:04d}"

        row = conn.execute(
            """
            insert into offers(user_id, user_name, client_name, offer_year, offer_seq, offer_no, status, vat_rate, valid_until, archived)
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, false)
            returning id
            """,
            (user_id, username_l, client_name, year, next_seq, offer_no, "DRAFT", 0, (date.today() + timedelta(days=14))),
        ).fetchone()
        return int(row["id"])


def get_offer(user_id: int, username: str, offer_id: int):
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        return conn.execute(
            f"""
            select o.id, o.user_id, o.user_name, o.client_name, o.created_at, o.offer_no, o.offer_year, o.offer_seq,
                   o.status, o.accepted_at, o.sent_at, o.archived, o.archived_at, o.client_email, o.client_address, o.client_oib,
                   o.terms_delivery, o.terms_payment, o.note, o.place, o.signed_by, o.vat_rate
            from offers o
            where o.id=%s and {_offer_owner_clause()}
            """,
            (offer_id, user_id, username_l),
        ).fetchone()


def _ensure_editable(offer_row: dict) -> None:
    if not offer_row:
        return
    if offer_row.get("archived"):
        raise ValueError("Ova ponuda je arhivirana.")
    if offer_row.get("status") == "ACCEPTED":
        raise ValueError("Ova ponuda je zaključana (ACCEPTED).")



def update_offer_client_email(user_id: int, username: str, offer_id: int, client_email: str | None) -> None:
    username_l = (username or "").strip().lower()
    email = (client_email or "").strip() or None
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute("update offers set client_email=%s where id=%s", (email, offer_id))


def update_offer_client_name(user_id: int, username: str, offer_id: int, client_name: str | None) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute(
            "update offers set client_name=%s where id=%s",
            (client_name, offer_id),
        )


def update_offer_client_details(
    user_id: int,
    username: str,
    offer_id: int,
    client_name: str | None,
    client_email: str | None,
    client_address: str | None,
    client_oib: str | None,
) -> None:
    username_l = (username or "").strip().lower()
    nm = (client_name or "").strip() or None
    em = (client_email or "").strip() or None
    addr = (client_address or "").strip() or None
    oib = (client_oib or "").strip() or None
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute(
            "update offers set client_name=%s, client_email=%s, client_address=%s, client_oib=%s where id=%s",
            (nm, em, addr, oib, offer_id),
        )

def update_offer_client_email(user_id: int, username: str, offer_id: int, client_email: str | None) -> None:
    username_l = (username or "").strip().lower()
    client_email = (client_email or "").strip() or None
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute(
            "update offers set client_email=%s where id=%s",
            (client_email, offer_id),
        )


def update_offer_meta(
    user_id: int,
    username: str,
    offer_id: int,
    terms_delivery: str | None,
    terms_payment: str | None,
    note: str | None,
    place: str | None,
    signed_by: str | None,
    vat_rate: float | None,
    valid_until: str | None,
) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute(
            """
            update offers
            set terms_delivery=%s,
                terms_payment=%s,
                note=%s,
                place=%s,
                signed_by=%s,
                vat_rate=%s,
                valid_until=%s
            where id=%s
            """,
            (terms_delivery, terms_payment, note, place, signed_by, float(vat_rate or 0), offer_id),
        )


def accept_offer(user_id: int, username: str, offer_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set status='ACCEPTED', accepted_at=now()
            where o.id=%s and {_offer_owner_clause()} and o.archived=false
            """,
            (offer_id, user_id, username_l),
        )


def unlock_offer(user_id: int, username: str, offer_id: int) -> None:
    """
    Safety valve: allow reverting ACCEPTED -> DRAFT (keeps numbering).
    """
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set status='DRAFT', accepted_at=null
            where o.id=%s and {_offer_owner_clause()} and o.archived=false
            """,
            (offer_id, user_id, username_l),
        )



def mark_offer_sent(user_id: int, username: str, offer_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set status='SENT', sent_at=now()
            where o.id=%s and {_offer_owner_clause()} and o.archived=false and o.status <> 'ACCEPTED'
            """,
            (offer_id, user_id, username_l),
        )


def record_email_result(user_id: int, username: str, offer_id: int, to_email: str, ok: bool, error: str | None = None) -> None:
    """Persist email send attempt outcome for audit/resend."""
    username_l = (username or "").strip().lower()
    to_email = (to_email or "").strip()
    err_txt = (error or "").strip() if not ok else None
    with get_conn() as conn:
        # bump attempts and store last result
        conn.execute(
            f"""
            update offers o
            set
              email_attempts = coalesce(o.email_attempts,0) + 1,
              last_email_to = %s,
              last_email_at = now(),
              last_email_error = %s,
              -- Mark as SENT on success (do not override ACCEPTED)
              status = case
                        when %s and o.status <> 'ACCEPTED' then 'SENT'
                        else o.status
                      end,
              sent_at = case
                        when %s and o.status <> 'ACCEPTED' then coalesce(o.sent_at, now())
                        else o.sent_at
                      end
            where o.id=%s and {_offer_owner_clause()} and o.archived=false
            """,
            (to_email, err_txt, ok, ok, offer_id, user_id, username_l),
        )


def duplicate_offer(user_id: int, username: str, offer_id: int) -> int:
    """Create a new DRAFT offer by copying meta + items from an existing offer."""
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        off = conn.execute(
            f"""
            select *
            from offers o
            where o.id=%s and {_offer_owner_clause()}
            """,
            (offer_id, user_id, username_l),
        ).fetchone()
        if not off:
            raise ValueError("Offer not found")

        # New offer gets a fresh number for current year
        new_id = create_offer(user_id, username_l, off.get("client_name"))
        # copy fields (keep DRAFT)
        conn.execute(
            f"""
            update offers o
            set
              client_email=%s,
              terms_delivery=%s,
              terms_payment=%s,
              note=%s,
              place=%s,
              signed_by=%s
            where o.id=%s and {_offer_owner_clause()}
            """,
            (
                off.get("client_email"),
                off.get("terms_delivery"),
                off.get("terms_payment"),
                off.get("note"),
                off.get("place"),
                off.get("signed_by"),
                new_id,
                user_id,
                username_l,
            ),
        )

        items = conn.execute(
            "select name,qty,price,line_total from offer_items where offer_id=%s order by id asc",
            (offer_id,),
        ).fetchall()
        for it in items:
            conn.execute(
                "insert into offer_items(offer_id,name,qty,price,line_total) values (%s,%s,%s,%s,%s)",
                (new_id, it["name"], it["qty"], it["price"], it["line_total"]),
            )
        return int(new_id)


def import_user_backup(user_id: int, username: str, payload: dict, restore_as_archived: bool = True) -> dict:
    """Import backup payload (offers.json structure). Returns stats."""
    username_l = (username or "").strip().lower()
    offers = payload.get("offers") or []
    items_map = payload.get("items") or {}
    imported = 0
    with get_conn() as conn:
        for o in offers:
            try:
                # create offer preserving number if present (no uniqueness constraint)
                row = conn.execute(
                    """
                    insert into offers(
                      user_id, user_name,
                      client_name, client_email,
                      created_at, sent_at,
                      offer_year, offer_seq, offer_no,
                      status, accepted_at,
                      archived, archived_at,
                      terms_delivery, terms_payment, note, place, signed_by,
                      vat_rate
                    )
                    values (
                      %s,%s,
                      %s,%s,
                      coalesce(%s::timestamptz, now()),
                      %s::timestamptz,
                      %s,%s,%s,
                      %s,
                      %s::timestamptz,
                      %s,
                      %s::timestamptz,
                      %s,%s,%s,%s,%s,
                      %s
                    )
                    returning id
                    """,
                    (
                        user_id, username_l,
                        o.get("client_name"), o.get("client_email"),
                        o.get("created_at"), o.get("sent_at"),
                        o.get("offer_year"), o.get("offer_seq"), o.get("offer_no"),
                        (o.get("status") or "DRAFT"),
                        o.get("accepted_at"),
                        bool(o.get("archived")) if not restore_as_archived else True,
                        o.get("archived_at"),
                        o.get("terms_delivery"), o.get("terms_payment"), o.get("note"), o.get("place"), o.get("signed_by"),
                        float(o.get("vat_rate") or 0),
                    ),
                ).fetchone()
                new_oid = int(row["id"])
                src_oid = str(o.get("id") or "")
                src_items = items_map.get(src_oid) or items_map.get(int(src_oid)) if src_oid.isdigit() else []
                for it in src_items:
                    conn.execute(
                        "insert into offer_items(offer_id,name,qty,price,line_total) values (%s,%s,%s,%s,%s)",
                        (new_oid, it.get("name"), float(it.get("qty") or 1), float(it.get("price") or 0), float(it.get("line_total") or 0)),
                    )
                imported += 1
            except Exception:
                # keep going on a single bad row
                continue
    return {"imported": imported}


def archive_offer(user_id: int, username: str, offer_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set archived=true, archived_at=now()
            where o.id=%s and {_offer_owner_clause()} and archived=false
            """,
            (offer_id, user_id, username_l),
        )


def unarchive_offer(user_id: int, username: str, offer_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set archived=false, archived_at=null
            where o.id=%s and {_offer_owner_clause()} and archived=true
            """,
            (offer_id, user_id, username_l),
        )


def delete_offer_permanently(user_id: int, username: str, offer_id: int) -> None:
    """
    Hard delete allowed only if archived=true (safety).
    """
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            f"select o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not row or not bool(row.get("archived")):
            return
        conn.execute("delete from offers where id=%s", (offer_id,))


# -----------------------------
# Items
# -----------------------------

def add_item(user_id: int, username: str, offer_id: int, name: str, qty: float, price: float) -> None:
    username_l = (username or "").strip().lower()
    q = float(qty or 0)
    p = float(price or 0)
    line_total = q * p
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute(
            """
            insert into offer_items(offer_id, name, qty, price, line_total)
            values (%s, %s, %s, %s, %s)
            """,
            (offer_id, (name or "").strip(), q, p, line_total),
        )


def list_items(offer_id: int):
    with get_conn() as conn:
        return conn.execute(
            """
            select id, name, qty, price, line_total
            from offer_items
            where offer_id=%s
            order by id asc
            """,
            (offer_id,),
        ).fetchall()


def delete_item(user_id: int, username: str, offer_id: int, item_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute("delete from offer_items where offer_id=%s and id=%s", (offer_id, item_id))


def clear_items(user_id: int, username: str, offer_id: int) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        offer = conn.execute(
            f"select o.status, o.archived from offers o where o.id=%s and {_offer_owner_clause()}",
            (offer_id, user_id, username_l),
        ).fetchone()
        if not offer:
            return
        _ensure_editable(dict(offer))
        conn.execute("delete from offer_items where offer_id=%s", (offer_id,))


# -----------------------------
# Lists
# -----------------------------

def list_offers(user_id: int, username: str, status: str | None = None, client_id: int | None = None, **kwargs) -> list[dict]:
    """List offers for the admin UI.

    This app is admin-only, and older versions wrote offers with various owner fields (user_id/user_name may be null).
    To avoid hiding existing data, we default to showing ALL offers, then apply optional filters.
    """
    # Backward-compatible params (older main.py used show/invoice/paid/client/q)
    show = (kwargs.get("show") or "").strip().lower()  # active|archived|all|""
    client = kwargs.get("client")
    q = (kwargs.get("q") or "").strip()
    try:
        if client_id is None and client is not None and str(client).strip():
            client_id = int(str(client).strip())
    except Exception:
        pass

    # status can be "ALL" from UI; treat as no filter
    if status and str(status).strip().upper() == "ALL":
        status = None

    where = ["1=1"]
    params: list = []

    # show filter (column archived exists in our schema; coalesce handles nulls)
    if show in ("active", ""):
        where.append("coalesce(archived,0)=0")
    elif show in ("archived",):
        where.append("coalesce(archived,0)=1")
    # else "all" -> no archived filter

    if status:
        where.append("status=%s")
        params.append(status)

    if client_id is not None:
        where.append("client_id=%s")
        params.append(client_id)

    if q:
        where.append("(offer_no ilike %s or client_name ilike %s)")
        like = f"%{q}%"
        params.extend([like, like])

    sql = "select * from offers where " + " and ".join(where) + " order by created_at desc"
    with get_conn() as conn:
        rows = conn.execute(sql, tuple(params)).fetchall()
    return [dict(r) for r in rows]


def list_clients_full(user_id: int, username: str) -> List[Dict[str, Any]]:
    """Return full client records for a user."""
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        rows = conn.execute(
            """
            select id, name, email, address, oib, note
            from clients
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
            order by name asc
            """,
            (user_id, username_l),
        ).fetchall()
        return [dict(r) for r in rows]


def list_clients(user_id: int, username: str):
    """Legacy helper returning only names."""
    return [(r.get("name"),) for r in list_clients_full(user_id, username)]


def get_client_by_name(user_id: int, username: str, name: str) -> Optional[Dict[str, Any]]:
    username_l = (username or "").strip().lower()
    nm = (name or "").strip()
    if not nm:
        return None
    with get_conn() as conn:
        row = conn.execute(
            """
            select id, name, email, address, oib, note
            from clients
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
              and name=%s
            limit 1
            """,
            (user_id, username_l, nm),
        ).fetchone()
        return dict(row) if row else None


def upsert_client_full(
    user_id: int,
    username: str,
    name: str,
    email: str | None = None,
    address: str | None = None,
    oib: str | None = None,
    note: str | None = None,
) -> None:
    """Insert/update a client by (user_name,name) unique key."""
    username_l = (username or "").strip().lower()
    nm = (name or "").strip()
    if not nm:
        return

    em = (email or "").strip() or None
    addr = (address or "").strip() or None
    o = (oib or "").strip() or None
    nt = (note or "").strip() or None

    with get_conn() as conn:
        conn.execute(
            """
            insert into clients(user_id, user_name, name, email, address, oib, note)
            values (%s, %s, %s, %s, %s, %s, %s)
            on conflict (user_name, name)
            do update set
              user_id = excluded.user_id,
              email = coalesce(excluded.email, clients.email),
              address = coalesce(excluded.address, clients.address),
              oib = coalesce(excluded.oib, clients.oib),
              note = coalesce(excluded.note, clients.note)
            """,
            (user_id, username_l, nm, em, addr, o, nt),
        )


def upsert_client(user_id: int, username: str, name: str) -> None:
    """Compatibility wrapper."""
    upsert_client_full(user_id, username, name)

def dashboard_monthly(user_id: int, username: str, year: int | None = None):
    """Return monthly counts and totals (subtotal) for a given year."""
    username_l = (username or "").strip().lower()
    yr = int(year or datetime.now().year)
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            select
              extract(month from o.created_at)::int as month,
              count(distinct o.id)::int as offers_count,
              coalesce(sum(i.line_total),0)::double precision as subtotal
            from offers o
            left join offer_items i on i.offer_id=o.id
            where ({_offer_owner_clause()})
              and extract(year from o.created_at)=%s
              and o.archived=false
            group by extract(month from o.created_at)
            order by month asc
            """,
            (user_id, username_l, yr),
        ).fetchall()
        # fill missing months
        bym = {int(r["month"]): {"month": int(r["month"]), "offers_count": int(r["offers_count"]), "subtotal": float(r["subtotal"])} for r in rows}
        out=[]
        for m in range(1,13):
            out.append(bym.get(m, {"month": m, "offers_count": 0, "subtotal": 0.0}))
        return out


# -----------------------------
# Settings
# -----------------------------

def get_settings(user_id: int, username: str) -> dict:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            """
            select user_name, user_id, company_name, company_address, company_oib, company_iban,
                   company_email, company_phone, logo_path,
                   (logo_bytes is not null) as has_logo,
                   logo_mime, logo_filename
            from company_settings
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
            """,
            (user_id, username_l),
        ).fetchone()
        return dict(row) if row else {}


def upsert_settings(
    user_id: int,
    username: str,
    company_name: str | None,
    company_address: str | None,
    company_oib: str | None,
    company_iban: str | None,
    company_email: str | None,
    company_phone: str | None,
    logo_path: str | None,
    logo_bytes: bytes | None = None,
    logo_mime: str | None = None,
    logo_filename: str | None = None,
) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            """
            insert into company_settings(
              user_name, user_id,
              company_name, company_address, company_oib, company_iban, company_email, company_phone, logo_path,
              logo_bytes, logo_mime, logo_filename
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            on conflict (user_name) do update set
              user_id=excluded.user_id,
              company_name=excluded.company_name,
              company_address=excluded.company_address,
              company_oib=excluded.company_oib,
              company_iban=excluded.company_iban,
              company_email=excluded.company_email,
              company_phone=excluded.company_phone,
              logo_path=excluded.logo_path,
              logo_bytes=coalesce(excluded.logo_bytes, company_settings.logo_bytes),
              logo_mime=coalesce(excluded.logo_mime, company_settings.logo_mime),
              logo_filename=coalesce(excluded.logo_filename, company_settings.logo_filename),
              updated_at=now()
            """,
            (
                username_l,
                int(user_id),
                company_name,
                company_address,
                company_oib,
                company_iban,
                company_email,
                company_phone,
                logo_path,
                psycopg.Binary(logo_bytes) if logo_bytes else None,
                logo_mime,
                logo_filename,
            ),
        )


def clear_logo(user_id: int, username: str) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            """
            update company_settings
            set logo_bytes=null, logo_mime=null, logo_filename=null
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
            """,
            (user_id, username_l),
        )


def get_logo_bytes(user_id: int, username: str) -> tuple[bytes | None, str | None]:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            """
            select logo_bytes, logo_mime
            from company_settings
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
            """,
            (user_id, username_l),
        ).fetchone()
        if not row:
            return None, None
        b = row.get("logo_bytes")
        # psycopg returns memoryview for bytea
        if b is not None and isinstance(b, memoryview):
            b = b.tobytes()
        return (b if b else None), (row.get("logo_mime") if row.get("logo_mime") else None)



# -----------------------------
# Audit / Admin
# -----------------------------

def log_audit(user_id: int | None, username: str | None, action: str, offer_id: int | None = None, ip: str | None = None, meta: Dict[str, Any] | None = None) -> None:
    try:
        with get_conn() as conn:
            conn.execute(
                "insert into audit_log(user_id, username, action, offer_id, ip, meta) values (%s,%s,%s,%s,%s,%s)",
                (user_id, (username or None), action, offer_id, (ip or None), json.dumps(meta or {}) if meta is not None else None),
            )
    except Exception:
        return


def list_audit(limit: int = 200) -> List[Dict[str, Any]]:
    with get_conn() as conn:
        rows = conn.execute(
            "select id, created_at, username, action, offer_id, ip, meta from audit_log order by id desc limit %s",
            (int(limit),),
        ).fetchall()
        return [dict(r) for r in rows]


# -----------------------------
# Templates (stored in company_settings)
# -----------------------------

DEFAULT_EMAIL_SUBJECT = "Ponuda {offer_no} – {client_name}"
DEFAULT_EMAIL_TEXT = "Poštovani,\n\nU prilogu je ponuda {offer_no}.\n\nLijep pozdrav,\n{company_name}"
DEFAULT_EMAIL_HTML = "<p>Poštovani,</p><p>U prilogu je ponuda <b>{offer_no}</b>.</p><p>Lijep pozdrav,<br>{company_name}</p>"
DEFAULT_PDF_FOOTER = ""

def get_templates(user_id: int, username: str) -> Dict[str, str]:
    s = get_settings(user_id, username) or {}
    return {
        "email_subject_tpl": s.get("email_subject_tpl") or DEFAULT_EMAIL_SUBJECT,
        "email_text_tpl": s.get("email_text_tpl") or DEFAULT_EMAIL_TEXT,
        "email_html_tpl": s.get("email_html_tpl") or DEFAULT_EMAIL_HTML,
        "pdf_footer_tpl": s.get("pdf_footer_tpl") or DEFAULT_PDF_FOOTER,
    }


def set_templates(user_id: int, username: str, subject: str, text_t: str, html_t: str, footer: str) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            """
            insert into company_settings(user_name, user_id, email_subject_tpl, email_text_tpl, email_html_tpl, pdf_footer_tpl, updated_at)
            values (%s, %s, %s, %s, %s, %s, now())
            on conflict (user_name) do update
            set user_id=excluded.user_id,
                email_subject_tpl=excluded.email_subject_tpl,
                email_text_tpl=excluded.email_text_tpl,
                email_html_tpl=excluded.email_html_tpl,
                pdf_footer_tpl=excluded.pdf_footer_tpl,
                updated_at=now()
            """,
            (username_l, user_id, subject, text_t, html_t, footer),
        )

# -----------------------------
# PDF / Excel exports
# -----------------------------

def _register_font(static_dir: str) -> str:
    font_path = Path(static_dir) / "DejaVuSans.ttf"
    if font_path.exists():
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", str(font_path)))
            return "DejaVuSans"
        except Exception:
            pass
    return "Helvetica"


def _draw_logo(c: canvas.Canvas, logo_bytes: bytes, x: float, y: float, max_w: float, max_h: float) -> float:
    """
    Draw logo with aspect ratio. Returns used height.
    """
    try:
        img = ImageReader(io.BytesIO(logo_bytes))
        iw, ih = img.getSize()
        if not iw or not ih:
            return 0.0
        scale = min(max_w / float(iw), max_h / float(ih))
        w = float(iw) * scale
        h = float(ih) * scale
        c.drawImage(img, x, y - h, width=w, height=h, mask="auto")
        return h
    except Exception:
        return 0.0



def _wrap_text(text: str, max_chars: int) -> list[str]:
    words = (text or "").replace("\r", "").split()
    lines: list[str] = []
    cur: list[str] = []
    cur_len = 0
    for w in words:
        if not cur:
            cur = [w]
            cur_len = len(w)
            continue
        if cur_len + 1 + len(w) <= max_chars:
            cur.append(w)
            cur_len += 1 + len(w)
        else:
            lines.append(" ".join(cur))
            cur = [w]
            cur_len = len(w)
    if cur:
        lines.append(" ".join(cur))
    return lines


def _draw_kv_block(c: canvas.Canvas, font: str, x: float, y: float, title: str, lines: list[str], width: float) -> float:
    """Draw a simple titled block and return new y (below the block)."""
    c.setFont(font, 10)
    c.rect(x, y - 6 - (14*len(lines)+18), width, 14*len(lines)+22, stroke=1, fill=0)
    c.setFont(font, 11)
    c.drawString(x + 8, y, title)
    yy = y - 18
    c.setFont(font, 10)
    for ln in lines:
        c.drawString(x + 8, yy, ln[:120])
        yy -= 14
    return yy - 10

def _draw_footer(c: canvas.Canvas, font: str, footer_text: str, x: float, y: float, w: float) -> float:
    """Draw footer text (wrapped). Returns height used."""
    if not footer_text:
        return 0.0
    c.setFont(font, 9)
    lines = []
    # preserve manual newlines
    for para in str(footer_text).split("\n"):
        para = para.strip()
        if not para:
            lines.append("")
            continue
        lines.extend(_wrap_text(para, max_chars=110))
    # draw from bottom up
    line_h = 11
    used = line_h * len(lines)
    yy = y
    for line in lines:
        c.drawString(x, yy, line)
        yy += line_h
    return used


import secrets

def ensure_public_token(user_id: int, username: str, offer_id: int) -> str:
    """Ensure offer has a stable public token for portal/tracking."""
    with get_conn() as conn:
        row = conn.execute(
            "select public_token from offers where id=%s and (user_id=%s or user_name=%s)",
            (offer_id, user_id, username),
        ).fetchone()
        if row and row.get("public_token"):
            return str(row["public_token"])
        token = secrets.token_urlsafe(18)
        conn.execute(
            "update offers set public_token=%s where id=%s and (user_id=%s or user_name=%s)",
            (token, offer_id, user_id, username),
        )
        return token

def get_offer_by_token(token: str) -> dict | None:
    with get_conn() as conn:
        row = conn.execute("select * from offers where public_token=%s", (token,)).fetchone()
        return dict(row) if row else None

def list_items_for_offer(offer_id: int) -> list[dict]:
    return list_items(offer_id)

def mark_offer_sent(user_id: int, username: str, offer_id: int) -> None:
    with get_conn() as conn:
        conn.execute(
            "update offers set status='SENT', sent_at=now() where id=%s and (user_id=%s or user_name=%s) and status!='ACCEPTED'",
            (offer_id, user_id, username),
        )

def track_open(token: str, ip: str | None = None) -> None:
    with get_conn() as conn:
        conn.execute(
            """
            update offers
               set view_count = coalesce(view_count,0)+1,
                   first_view_at = coalesce(first_view_at, now()),
                   last_view_at = now(),
                   last_view_ip = %s
             where public_token = %s
            """,
            (ip, token),
        )

def track_click(token: str, ip: str | None = None) -> None:
    with get_conn() as conn:
        conn.execute(
            """
            update offers
               set click_count = coalesce(click_count,0)+1,
                   last_click_at = now(),
                   last_click_ip = %s
             where public_token = %s
            """,
            (ip, token),
        )

def accept_by_token(token: str, ip: str | None = None) -> bool:
    with get_conn() as conn:
        res = conn.execute(
            """
            update offers
               set status='ACCEPTED',
                   accepted_at=now(),
                   accepted_via='PORTAL',
                   accepted_via_at=now()
             where public_token=%s and status!='ACCEPTED'
            """,
            (token,),
        )
        return res.rowcount > 0

def render_offer_pdf(
    offer: Dict[str, Any],
    items: List[Dict[str, Any]],
    settings: Dict[str, Any],
    static_dir: str,
    logo_bytes: bytes | None = None,
) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    font = _register_font(static_dir)

    top_y = h - 40
    logo_h = 0.0
    if logo_bytes:
        logo_h = _draw_logo(c, logo_bytes, x=40, y=top_y, max_w=140, max_h=60)

    c.setFont(font, 16)
    c.drawString(40 + (160 if logo_h else 0), h - 50, "Ponuda")

    c.setFont(font, 10)
    y_company = h - 70
    x_company = 40 + (160 if logo_h else 0)
    if settings.get("company_name"):
        c.drawString(x_company, y_company, str(settings.get("company_name")))
        y_company -= 14
    if settings.get("company_address"):
        c.drawString(x_company, y_company, str(settings.get("company_address")))
        y_company -= 14

    if settings.get("company_oib"):
        c.drawString(x_company, y_company, f"OIB: {settings.get('company_oib')}")
        y_company -= 14
    if settings.get("company_iban"):
        c.drawString(x_company, y_company, f"IBAN: {settings.get('company_iban')}")
        y_company -= 14
    if settings.get("company_email"):
        c.drawString(x_company, y_company, f"E-mail: {settings.get('company_email')}")
        y_company -= 14
    if settings.get("company_phone"):
        c.drawString(x_company, y_company, f"Tel: {settings.get('company_phone')}")
        y_company -= 14

    if settings.get("company_oib"):
        c.drawString(x_company, y_company, f"OIB: {settings.get('company_oib')}")
        y_company -= 14
    if settings.get("company_iban"):
        c.drawString(x_company, y_company, f"IBAN: {settings.get('company_iban')}")
        y_company -= 14
    if settings.get("company_email"):
        c.drawString(x_company, y_company, f"E-mail: {settings.get('company_email')}")
        y_company -= 14
    if settings.get("company_phone"):
        c.drawString(x_company, y_company, f"Tel: {settings.get('company_phone')}")
        y_company -= 14

    c.drawRightString(w - 40, h - 70, f"Broj: {offer.get('offer_no') or ''}")
    c.drawRightString(w - 40, h - 85, f"Datum: {str(offer.get('created_at') or '')[:16]}")
    if offer.get("status") == "ACCEPTED":
        c.drawRightString(w - 40, h - 100, "Status: ACCEPTED")

    if offer.get("invoice_no"):
        c.drawRightString(w - 40, h - 115, f"Račun: {offer.get('invoice_no')}")

    # Client block
    client_lines = []
    if offer.get("client_name"):
        client_lines.append(str(offer.get("client_name")))
    if offer.get("client_address"):
        for ln in _wrap_text(str(offer.get("client_address")), 52):
            client_lines.append(ln)
    if offer.get("client_oib"):
        client_lines.append(f"OIB: {offer.get('client_oib')}")
    if offer.get("client_email"):
        client_lines.append(f"E-mail: {offer.get('client_email')}")

    y_after_client = _draw_kv_block(c, font, x=40, y=h - 132, title="Kupac", lines=client_lines or [""], width=w - 80)



    # Table header
    y = y_after_client
    c.setFont(font, 10)
    c.drawString(40, y, "Naziv")
    c.drawRightString(w - 220, y, "Količina")
    c.drawRightString(w - 140, y, "Cijena")
    c.drawRightString(w - 40, y, "Ukupno")
    y -= 10
    c.line(40, y, w - 40, y)
    y -= 16

    subtotal = Decimal("0")
    for it in items:
        name = str(it.get("name") or "")
        qty = Decimal(str(it.get("qty", 0) or 0))
        price = Decimal(str(it.get("price", 0) or 0))
        line_total = qty * price
        subtotal += line_total

        c.drawString(40, y, name[:60])
        c.drawRightString(w - 220, y, f"{qty:.2f}")
        c.drawRightString(w - 140, y, f"{price:.2f}")
        c.drawRightString(w - 40, y, f"{line_total:.2f}")
        y -= 14
        if y < 90:
            c.showPage()
            c.setFont(font, 10)
            y = h - 60

    vat_rate = Decimal(str(offer.get("vat_rate", 0) or 0))
    vat = (subtotal * vat_rate / Decimal("100")) if vat_rate else Decimal("0")
    total = subtotal + vat

    y -= 8
    c.line(40, y, w - 40, y)
    y -= 18
    c.drawRightString(w - 40, y, f"Međuzbroj: {subtotal:.2f} €")
    y -= 14
    if vat_rate and vat_rate != Decimal("0"):
        c.drawRightString(w - 40, y, f"PDV {vat_rate:.0f}%: {vat:.2f} €")
        y -= 16
    else:
        y -= 2
    c.setFont(font, 12)
    c.drawRightString(w - 40, y, f"Ukupno: {total:.2f} €")

    # Meta lines
    c.setFont(font, 10)
    y -= 28
    for label, key in [
        ("Mjesto", "place"),
        ("Rok isporuke", "terms_delivery"),
        ("Rok plaćanja", "terms_payment"),
        ("Napomena", "note"),
        ("Potpis", "signed_by"),
    ]:
        val = offer.get(key) or ""
        if val:
            c.drawString(40, y, f"{label}: {val}")
            y -= 14

    c.showPage()
    
    # Footer (template)
    footer_tpl = (settings.get("pdf_footer_tpl") or "").strip()
    if footer_tpl:
        _draw_footer(c, font, footer_tpl, x=40, y=28, w=w-80)
    c.save()
    return buf.getvalue()


def render_invoice_pdf(
    offer: Dict[str, Any],
    items: List[Dict[str, Any]],
    settings: Dict[str, Any],
    static_dir: str,
    logo_bytes: bytes | None = None,
) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    font = _register_font(static_dir)

    top_y = h - 40
    logo_h = 0.0
    if logo_bytes:
        logo_h = _draw_logo(c, logo_bytes, x=40, y=top_y, max_w=140, max_h=60)

    c.setFont(font, 16)
    c.drawString(40 + (160 if logo_h else 0), h - 50, "Račun")

    c.setFont(font, 10)
    y_company = h - 70
    x_company = 40 + (160 if logo_h else 0)
    if settings.get("company_name"):
        c.drawString(x_company, y_company, str(settings.get("company_name")))
        y_company -= 14
    if settings.get("company_address"):
        c.drawString(x_company, y_company, str(settings.get("company_address")))
        y_company -= 14

    if settings.get("company_oib"):
        c.drawString(x_company, y_company, f"OIB: {settings.get('company_oib')}")
        y_company -= 14
    if settings.get("company_iban"):
        c.drawString(x_company, y_company, f"IBAN: {settings.get('company_iban')}")
        y_company -= 14
    if settings.get("company_email"):
        c.drawString(x_company, y_company, f"E-mail: {settings.get('company_email')}")
        y_company -= 14
    if settings.get("company_phone"):
        c.drawString(x_company, y_company, f"Tel: {settings.get('company_phone')}")
        y_company -= 14

    if settings.get("company_oib"):
        c.drawString(x_company, y_company, f"OIB: {settings.get('company_oib')}")
        y_company -= 14
    if settings.get("company_iban"):
        c.drawString(x_company, y_company, f"IBAN: {settings.get('company_iban')}")
        y_company -= 14
    if settings.get("company_email"):
        c.drawString(x_company, y_company, f"E-mail: {settings.get('company_email')}")
        y_company -= 14
    if settings.get("company_phone"):
        c.drawString(x_company, y_company, f"Tel: {settings.get('company_phone')}")
        y_company -= 14

    c.drawRightString(w - 40, h - 70, f"Račun broj: {offer.get('invoice_no') or ''}")
    inv_date = offer.get("invoice_date") or offer.get("accepted_at") or offer.get("created_at") or ""
    c.drawRightString(w - 40, h - 85, f"Datum: {str(inv_date)[:16]}")
    c.drawRightString(w - 40, h - 100, f"Ponuda: {offer.get('offer_no') or ''}")
    c.drawRightString(w - 40, h - 115, f"Status: {'PLAĆENO' if offer.get('paid') else 'NIJE PLAĆENO'}")

    # Client block
    client_lines = []
    if offer.get("client_name"):
        client_lines.append(str(offer.get("client_name")))
    if offer.get("client_address"):
        for ln in _wrap_text(str(offer.get("client_address")), 52):
            client_lines.append(ln)
    if offer.get("client_oib"):
        client_lines.append(f"OIB: {offer.get('client_oib')}")
    if offer.get("client_email"):
        client_lines.append(f"E-mail: {offer.get('client_email')}")

    y_after_client = _draw_kv_block(c, font, x=40, y=h - 142, title="Kupac", lines=client_lines or [""], width=w - 80)


    if offer.get("client_email"):
        c.drawString(40, h - 145, f"E-mail: {offer.get('client_email') or ''}")

    y = y_after_client
    c.setFont(font, 10)
    c.drawString(40, y, "Naziv")
    c.drawRightString(w - 220, y, "Količina")
    c.drawRightString(w - 140, y, "Cijena")
    c.drawRightString(w - 40, y, "Ukupno")
    y -= 10
    c.line(40, y, w - 40, y)
    y -= 16

    subtotal = Decimal("0")
    for it in items:
        name = str(it.get("name") or "")
        qty = Decimal(str(it.get("qty", 0) or 0))
        price = Decimal(str(it.get("price", 0) or 0))
        line_total = qty * price
        subtotal += line_total

        c.drawString(40, y, name[:60])
        c.drawRightString(w - 220, y, f"{qty:.2f}")
        c.drawRightString(w - 140, y, f"{price:.2f}")
        c.drawRightString(w - 40, y, f"{line_total:.2f}")
        y -= 14
        if y < 110:
            c.showPage()
            c.setFont(font, 10)
            y = h - 60

    vat_rate = Decimal("25")
    vat = subtotal * vat_rate / Decimal("100")
    total = subtotal + vat

    y -= 8
    c.line(40, y, w - 40, y)
    y -= 18
    c.drawRightString(w - 40, y, f"Međuzbroj: {subtotal:.2f} €")
    y -= 14
    if vat_rate and vat_rate != Decimal("0"):
        c.drawRightString(w - 40, y, f"PDV {vat_rate:.0f}%: {vat:.2f} €")
        y -= 16
    else:
        y -= 2
    c.setFont(font, 12)
    c.drawRightString(w - 40, y, f"Ukupno: {total:.2f} €")

    footer = (settings.get("pdf_footer_tpl") or "").strip()
    if footer:
        c.setFont(font, 9)
        c.drawString(40, 40, footer[:140])

    c.showPage()
    
    # Footer (template)
    footer_tpl = (settings.get("pdf_footer_tpl") or "").strip()
    if footer_tpl:
        _draw_footer(c, font, footer_tpl, x=40, y=28, w=w-80)
    c.save()
    return buf.getvalue()


def render_offer_excel(offer: Dict[str, Any], items: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ponuda"

    ws.append(["Naziv", "Količina", "Cijena", "Ukupno"])
    for it in items:
        qty = float(it.get("qty", 0) or 0)
        price = float(it.get("price", 0) or 0)
        ws.append([it.get("name", ""), qty, price, qty * price])

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def export_user_backup_zip(user_id: int, username: str, static_dir: str) -> bytes:
    """Create a ZIP: offers.json + PDFs for all non-archived offers of the user."""
    import zipfile
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        offers = conn.execute(
            f"""
            select o.*
            from offers o
            where ({_offer_owner_clause()})
            order by o.created_at desc, o.id desc
            """,
            (user_id, username_l),
        ).fetchall()

        # collect items per offer
        offers_out = []
        items_map = {}
        for o in offers:
            oid = int(o["id"])
            items = conn.execute(
                "select id,name,qty,price,line_total from offer_items where offer_id=%s order by id asc",
                (oid,),
            ).fetchall()
            items_map[oid] = [dict(it) for it in items]
            offers_out.append(dict(o))

    # Build zip
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("offers.json", json.dumps({"user": username_l, "exported_at": datetime.now().isoformat(), "offers": offers_out, "items": items_map}, ensure_ascii=False, indent=2))
        # PDFs
        settings = get_settings(user_id, username_l)
        logo_bytes, _ = get_logo_bytes(user_id, username_l)
        for o in offers_out:
            oid = int(o["id"])
            pdf = render_offer_pdf(offer=o, items=items_map.get(oid, []), settings=settings, static_dir=static_dir, logo_bytes=logo_bytes)
            offer_no = (o.get("offer_no") or str(oid)).replace("/", "-")
            z.writestr(f"pdf/ponuda_{offer_no}.pdf", pdf)
    return buf.getvalue()
# -----------------------------
# Invoices (stored on offers rows)
# -----------------------------

def create_invoice_from_offer(user_id: int, username: str, offer_id: int) -> Dict[str, Any]:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        row = conn.execute(
            f"select * from offers o where ({_offer_owner_clause()}) and o.id=%s",
            (user_id, username_l, offer_id),
        ).fetchone()
        if not row:
            raise ValueError("Offer not found")
        offer = dict(row)

        if offer.get("status") != "ACCEPTED":
            raise ValueError("Only ACCEPTED offers can be invoiced")
        if offer.get("invoice_no"):
            return offer

        year = datetime.now().year
        row_seq = conn.execute(
            """
            select coalesce(max(invoice_seq), 0) as max_seq
            from offers
            where (user_id=%s or (user_id is null and lower(user_name)=%s))
              and invoice_year=%s
            """,
            (user_id, username_l, year),
        ).fetchone()
        next_seq = int((row_seq or {}).get("max_seq") or 0) + 1
        invoice_no = f"{year}-{next_seq:04d}"

        row2 = conn.execute(
            """
            update offers
            set is_invoice=true,
                invoice_year=%s,
                invoice_seq=%s,
                invoice_no=%s,
                invoice_date=now()
            where id=%s
            returning *
            """,
            (year, next_seq, invoice_no, offer_id),
        ).fetchone()
        return dict(row2)


def set_invoice_paid(user_id: int, username: str, offer_id: int, paid: bool) -> None:
    username_l = (username or "").strip().lower()
    with get_conn() as conn:
        conn.execute(
            f"""
            update offers o
            set paid=%s,
                paid_at=case when %s then now() else null end
            where ({_offer_owner_clause()}) and o.id=%s and o.invoice_no is not null
            """,
            (bool(paid), bool(paid), user_id, username_l, offer_id),
        )


